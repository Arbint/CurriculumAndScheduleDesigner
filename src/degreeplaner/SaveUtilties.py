import json
import os
from degreeplaner.CourseList import CourseListModel
from degreeplaner.PathUtility import GetPrjDir
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def ConvertModelsToJson(models: list[CourseListModel]):
    data = {}
    for model in models:
        data[model.listName] = model.CoursesToJsonList()

    return data

def LoadJsonDictFromPath(jsonPath)->dict:
    with open(jsonPath, 'r') as jsonFile:
        data = json.load(jsonFile)
        return data

def SaveModelsToJson(models: list[CourseListModel], savePath):
    data = ConvertModelsToJson(models)
    SaveToJson(data, savePath)

def SaveToJson(jsonDict, savePath):
    with open(savePath, "w") as jsonFile:
        json.dump(jsonDict, jsonFile, indent=4)

def GetDefaultSaveDir():
    return os.path.join(GetPrjDir(), "saves")

SKIP_MODELS = {"All Classes"}

def ExportModelsToExcel(models, savePath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Degree Plan"

    bold = Font(bold=True)
    headerFill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    grandTotal = 0

    for model in models:
        if model.listName in SKIP_MODELS:
            continue
        if not model.courses:
            continue

        ws.append([model.listName])
        headerRow = ws.max_row
        ws.merge_cells(f"A{headerRow}:D{headerRow}")
        cell = ws[f"A{headerRow}"]
        cell.font = bold
        cell.fill = headerFill
        cell.alignment = Alignment(horizontal="center")

        semesterCredits = 0
        for course in model.courses:
            credits = course.GetCredits()
            semesterCredits += credits
            row = [
                f"{course.departmentPrefix} {course.courseNumber}",
                course.courseName,
                f"{credits} credits",
            ]
            if course.note:
                row.append(course.note)
            ws.append(row)

        ws.append([f"Total: {semesterCredits} credits"])
        ws[f"A{ws.max_row}"].font = bold
        ws.append([])

        grandTotal += semesterCredits

    ws.append([f"Grand Total: {grandTotal} credits"])
    ws[f"A{ws.max_row}"].font = bold

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    wb.save(savePath)
